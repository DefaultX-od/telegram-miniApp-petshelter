import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from db_calls import get_pet_types, get_pet_statuses, get_pets

def generate_template():
    pet_types = get_pet_types()
    pet_statuses = get_pet_statuses()

    # Загружаем шаблон
    pets_template = openpyxl.load_workbook('output/pets-data-template.xlsx')
    working_sheet = pets_template.active

    for row in working_sheet['A2:H11']:
        for cell in row:
            cell.protection = openpyxl.styles.Protection(locked=False)

    # Создаем объекты DataValidation для колонок
    pet_type_validation = DataValidation(
        type="list", formula1=f'"{",".join(pet_types)}"', showDropDown=False
    )
    pet_status_validation = DataValidation(
        type="list", formula1=f'"{",".join(pet_statuses)}"', showDropDown=False
    )

    # Добавляем валидацию в нужные диапазоны
    pet_type_validation.add('A2:A11')  # Используйте add() вместо append
    pet_status_validation.add('B2:B11')  # Используйте add() вместо append

    # Применяем валидацию к ячейкам
    working_sheet.add_data_validation(pet_type_validation)
    working_sheet.add_data_validation(pet_status_validation)

    # Защищаем лист
    working_sheet.protection.sheet = True
    # working_sheet.protection.set_password('1234')  # Без пароля

    # Сохраняем файл
    pets_template.save('output/pets-data-template.xlsx')

def get_pets_data_from_file():
    pets_file = openpyxl.load_workbook('output/pets-data-file.xlsx')
    working_sheet = pets_file.active

    pets_data = []
    for row in working_sheet.iter_rows(min_row=2, min_col=1, max_row=working_sheet.max_row, max_col=working_sheet.max_column):
        row_data = [cell.value for cell in row]
        if any(cell is not None and cell != "" for cell in row_data):
            pets_data.append(row_data)

    return pets_data

def get_pets_table_as_file():
    pets = get_pets()
    pet_types = get_pet_types()
    pet_statuses = get_pet_statuses()

    pets_table_file = openpyxl.load_workbook('output/pets-table-template.xlsx')
    working_sheet = pets_table_file.active

    for row in working_sheet[f'A2:H{str(len(pets) + 1)}']:
        for cell in row:
            cell.protection = openpyxl.styles.Protection(locked=False)
    
    # Создаем объекты DataValidation для каждой колонки
    pet_type_validation = DataValidation(
        type="list", formula1=f'"{",".join(pet_types)}"', showDropDown=False
    )
    pet_status_validation = DataValidation(
        type="list", formula1=f'"{",".join(pet_statuses)}"', showDropDown=False
    )
    sex_validation = DataValidation(
        type="list", formula1='"Мужской,Женский"', showDropDown=False
    )
    fertility_validation = DataValidation(
        type="list", formula1='"Да,Нет"', showDropDown=False
    )

    # Применяем валидацию к диапазонам для каждой колонки
    pet_type_validation.add(f'B2:B{str(len(pets) + 1)}')  # Площадь для pet_type
    pet_status_validation.add(f'C2:C{str(len(pets) + 1)}')  # Площадь для pet_status
    sex_validation.add(f'E2:E{str(len(pets) + 1)}')  # Площадь для sex
    fertility_validation.add(f'H2:H{str(len(pets) + 1)}')  # Площадь для fertility

    # Добавляем валидацию в нужные диапазоны
    working_sheet.add_data_validation(pet_type_validation)
    working_sheet.add_data_validation(pet_status_validation)
    working_sheet.add_data_validation(sex_validation)
    working_sheet.add_data_validation(fertility_validation)

    row = 2

    # Заполнение данными после применения валидации
    for pet in pets:
        working_sheet.cell(row=row, column=1).value = str(pet["id"])
        working_sheet.cell(row=row, column=2).value = str(pet["type"])
        working_sheet.cell(row=row, column=3).value = str(pet["status"])
        working_sheet.cell(row=row, column=4).value = str(pet["name"])
        working_sheet.cell(row=row, column=5).value = str(pet["sex"])
        working_sheet.cell(row=row, column=6).value = str(pet["age"])
        working_sheet.cell(row=row, column=7).value = str(pet["img"]) if pet["img"] is not None else ""
        working_sheet.cell(row=row, column=8).value = str(pet["fertility"])
        working_sheet.cell(row=row, column=9).value = str(pet["description"])
        row += 1

    # Защищаем лист (по желанию)
    working_sheet.protection.sheet = True
    # working_sheet.protection.set_password('1234')  # Можно добавить пароль

    # Сохраняем файл с данными и валидацией
    pets_table_file.save('output/pets-table-file.xlsx')

